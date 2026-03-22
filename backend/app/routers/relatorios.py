"""
Router: Relatórios, Exportação, Importação e Dashboard de Temporalidade.

HU-032  Exportação PDF do CCD/TTD
HU-033  Exportação CSV/Excel do CCD/TTD
HU-034  Importação em lote (Excel)
HU-035  Dashboard prazos vencidos
HU-036  Busca/filtro avançado no PCD
HU-037  Listagem / Termo de Eliminação
HU-038  Relatório de Transferência / Recolhimento
"""

import csv
import io
import uuid
from datetime import datetime, UTC

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from pydantic import BaseModel, Field
from sqlalchemy import select, or_, func as sa_func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.pcd import PcdNivel
from app.models.ttd import RegraRetencao, LegalHold, OrdemDestinacao
from app.models.user import User
from app.routers.auth import get_current_user

router = APIRouter()

# ========== Constantes / tipos válidos ==========
TIPOS_NIVEL_VALIDOS = {"funcao", "subfuncao", "atividade", "serie", "classe", "tipo_documental"}
DESTINACOES_VALIDAS = {"eliminacao", "guarda_permanente", "microfilmagem", "amostragem"}

LABEL_TIPO = {
    "funcao": "Função",
    "subfuncao": "Subfunção",
    "atividade": "Atividade",
    "serie": "Série",
    "classe": "Classe",
    "tipo_documental": "Tipo Documental",
}

LABEL_DESTINACAO = {
    "eliminacao": "Eliminação",
    "guarda_permanente": "Guarda Permanente",
    "microfilmagem": "Microfilmagem",
    "amostragem": "Amostragem",
}


# ========== Schemas ==========

class BuscaPcdResponse(BaseModel):
    id: uuid.UUID
    pai_id: uuid.UUID | None
    tipo: str
    codigo: str
    titulo: str
    descricao: str | None
    codigo_conarq: str | None
    status: str
    nivel_sigilo: str
    model_config = {"from_attributes": True}


class PrazoVencidoItem(BaseModel):
    pcd_nivel_id: uuid.UUID
    codigo: str
    titulo: str
    tipo: str
    regra_id: uuid.UUID
    evento_inicio: str
    prazo_dias: int
    fase_corrente: int
    fase_intermediaria: int
    destinacao_final: str
    base_legal: str | None
    regra_criada_em: datetime
    tem_hold_ativo: bool


class DashboardTemporalidadeResponse(BaseModel):
    total_regras: int
    total_com_hold: int
    por_destinacao: dict[str, int]
    itens: list[PrazoVencidoItem]


class ImportResultResponse(BaseModel):
    criados: int
    erros: list[str]


class TermoEliminacaoItem(BaseModel):
    codigo: str
    titulo: str
    tipo: str
    destinacao_final: str
    base_legal: str | None
    fase_corrente: int
    fase_intermediaria: int
    observacoes: str | None


class TermoEliminacaoResponse(BaseModel):
    titulo: str
    data_geracao: datetime
    gerado_por: str
    itens: list[TermoEliminacaoItem]
    total_itens: int


class RelatorioTransferenciaItem(BaseModel):
    codigo: str
    titulo: str
    tipo: str
    destinacao_final: str
    fase_corrente: int
    fase_intermediaria: int
    base_legal: str | None


class RelatorioTransferenciaResponse(BaseModel):
    titulo: str
    tipo_relatorio: str
    data_geracao: datetime
    gerado_por: str
    itens: list[RelatorioTransferenciaItem]
    total_itens: int


# ========== Helpers ==========

def _flatten_tree(niveis: list[PcdNivel], depth: int = 0) -> list[tuple[PcdNivel, int]]:
    """Achatar árvore em lista com profundidade."""
    result: list[tuple[PcdNivel, int]] = []
    for n in sorted(niveis, key=lambda x: x.codigo):
        result.append((n, depth))
        if n.filhos:
            result.extend(_flatten_tree(n.filhos, depth + 1))
    return result


async def _carregar_arvore_completa(db: AsyncSession) -> list[PcdNivel]:
    result = await db.execute(
        select(PcdNivel)
        .where(PcdNivel.pai_id.is_(None))
        .options(
            selectinload(PcdNivel.filhos)
            .selectinload(PcdNivel.filhos)
            .selectinload(PcdNivel.filhos)
            .selectinload(PcdNivel.filhos)
        )
    )
    return list(result.scalars().all())


async def _carregar_regras_por_nivel(db: AsyncSession) -> dict[uuid.UUID, RegraRetencao]:
    result = await db.execute(select(RegraRetencao))
    regras = result.scalars().all()
    mapa: dict[uuid.UUID, RegraRetencao] = {}
    for r in regras:
        mapa[r.pcd_nivel_id] = r
    return mapa


async def _holds_ativos_por_nivel(db: AsyncSession) -> set[uuid.UUID]:
    result = await db.execute(
        select(LegalHold.pcd_nivel_id).where(LegalHold.status == "ativo").distinct()
    )
    return {row[0] for row in result.all()}


# ========== HU-036: Busca / Filtro Avançado ==========

@router.get("/busca", response_model=list[BuscaPcdResponse])
async def buscar_pcd(
    q: str | None = Query(None, description="Texto livre (código, título, descrição)"),
    tipo: str | None = Query(None, description="Filtrar por tipo"),
    nivel_sigilo: str | None = Query(None, description="Filtrar por sigilo"),
    status: str | None = Query(None, description="Filtrar por status"),
    codigo_conarq: str | None = Query(None, description="Filtrar por código CONARQ"),
    destinacao_final: str | None = Query(None, description="Filtrar por destinação final (via regras TTD)"),
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Busca avançada no PCD com filtros combinados (HU-036)."""
    query = select(PcdNivel)

    if q:
        pattern = f"%{q}%"
        query = query.where(
            or_(
                PcdNivel.codigo.ilike(pattern),
                PcdNivel.titulo.ilike(pattern),
                PcdNivel.descricao.ilike(pattern),
                PcdNivel.codigo_conarq.ilike(pattern),
            )
        )

    if tipo:
        query = query.where(PcdNivel.tipo == tipo)
    if nivel_sigilo:
        query = query.where(PcdNivel.nivel_sigilo == nivel_sigilo)
    if status:
        query = query.where(PcdNivel.status == status)
    if codigo_conarq:
        query = query.where(PcdNivel.codigo_conarq.ilike(f"%{codigo_conarq}%"))

    if destinacao_final:
        sub = select(RegraRetencao.pcd_nivel_id).where(
            RegraRetencao.destinacao_final == destinacao_final
        )
        query = query.where(PcdNivel.id.in_(sub))

    query = query.order_by(PcdNivel.codigo).limit(limit)
    result = await db.execute(query)
    return result.scalars().all()


# ========== HU-035: Dashboard Prazos Vencidos ==========

@router.get("/dashboard-temporalidade", response_model=DashboardTemporalidadeResponse)
async def dashboard_temporalidade(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Dashboard de temporalidade com indicadores de prazos (HU-035)."""
    regras = await db.execute(select(RegraRetencao).order_by(RegraRetencao.created_at.desc()))
    regras_list = regras.scalars().all()

    holds_ativos = await _holds_ativos_por_nivel(db)

    nivel_ids = {r.pcd_nivel_id for r in regras_list}
    niveis_result = await db.execute(select(PcdNivel).where(PcdNivel.id.in_(nivel_ids)))
    niveis_map = {n.id: n for n in niveis_result.scalars().all()}

    por_destinacao: dict[str, int] = {}
    itens: list[PrazoVencidoItem] = []

    for regra in regras_list:
        dest = regra.destinacao_final
        por_destinacao[dest] = por_destinacao.get(dest, 0) + 1

        nivel = niveis_map.get(regra.pcd_nivel_id)
        if not nivel:
            continue

        itens.append(PrazoVencidoItem(
            pcd_nivel_id=nivel.id,
            codigo=nivel.codigo,
            titulo=nivel.titulo,
            tipo=nivel.tipo,
            regra_id=regra.id,
            evento_inicio=regra.evento_inicio,
            prazo_dias=regra.prazo_dias,
            fase_corrente=regra.fase_corrente,
            fase_intermediaria=regra.fase_intermediaria,
            destinacao_final=regra.destinacao_final,
            base_legal=regra.base_legal,
            regra_criada_em=regra.created_at,
            tem_hold_ativo=regra.pcd_nivel_id in holds_ativos,
        ))

    return DashboardTemporalidadeResponse(
        total_regras=len(regras_list),
        total_com_hold=sum(1 for i in itens if i.tem_hold_ativo),
        por_destinacao=por_destinacao,
        itens=itens,
    )


# ========== HU-032: Exportação PDF ==========

@router.get("/exportar/pdf")
async def exportar_pdf(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exportar CCD/TTD completo em PDF tabular (HU-032)."""
    arvore = await _carregar_arvore_completa(db)
    flat = _flatten_tree(arvore)
    regras_map = await _carregar_regras_por_nivel(db)
    holds = await _holds_ativos_por_nivel(db)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Cabeçalho
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Codigo de Classificacao de Documentos e Tabela de Temporalidade", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Gerado em: {datetime.now(UTC).strftime('%d/%m/%Y %H:%M UTC')}  |  Por: {current_user.email}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(4)

    # Cabeçalho da tabela
    col_w = [15, 25, 30, 55, 25, 20, 20, 25, 35, 30]
    headers = ["Ind.", "Codigo", "CONARQ", "Titulo", "Tipo", "Corrente", "Intermed.", "Destinacao", "Base Legal", "Obs."]

    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(200, 200, 200)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 6)
    for idx, (nivel, depth) in enumerate(flat, 1):
        regra = regras_map.get(nivel.id)
        indent = "  " * depth

        vals = [
            str(idx),
            nivel.codigo,
            nivel.codigo_conarq or "",
            f"{indent}{nivel.titulo}"[:50],
            LABEL_TIPO.get(nivel.tipo, nivel.tipo),
            f"{regra.fase_corrente}a" if regra else "-",
            f"{regra.fase_intermediaria}a" if regra else "-",
            LABEL_DESTINACAO.get(regra.destinacao_final, "") if regra else "-",
            (regra.base_legal or "")[:30] if regra else "",
            (regra.observacoes or "")[:25] if regra else "",
        ]

        hold_ativo = nivel.id in holds
        if hold_ativo:
            pdf.set_fill_color(255, 255, 200)
        else:
            pdf.set_fill_color(255, 255, 255)

        for i, v in enumerate(vals):
            pdf.cell(col_w[i], 5, v, border=1, fill=hold_ativo)
        pdf.ln()

    # Legenda
    pdf.ln(3)
    pdf.set_font("Helvetica", "I", 7)
    pdf.cell(0, 5, "* Linhas amarelas indicam niveis com Legal Hold ativo (prazo suspenso)", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 5, "a = anos  |  Destinacao: Eliminacao, Guarda Permanente, Microfilmagem, Amostragem", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=CCD_TTD_Completo.pdf"},
    )


# ========== HU-033: Exportação CSV / Excel ==========

@router.get("/exportar/csv")
async def exportar_csv(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exportar CCD/TTD em CSV (HU-033)."""
    arvore = await _carregar_arvore_completa(db)
    flat = _flatten_tree(arvore)
    regras_map = await _carregar_regras_por_nivel(db)

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([
        "Código", "Código CONARQ", "Tipo", "Título", "Descrição",
        "Nível Sigilo", "Status", "Evento Início",
        "Fase Corrente (anos)", "Fase Intermediária (anos)",
        "Destinação Final", "Base Legal", "Legislação Ref", "Observações",
    ])

    for nivel, _depth in flat:
        regra = regras_map.get(nivel.id)
        writer.writerow([
            nivel.codigo,
            nivel.codigo_conarq or "",
            nivel.tipo,
            nivel.titulo,
            nivel.descricao or "",
            nivel.nivel_sigilo,
            nivel.status,
            regra.evento_inicio if regra else "",
            regra.fase_corrente if regra else "",
            regra.fase_intermediaria if regra else "",
            regra.destinacao_final if regra else "",
            regra.base_legal or "" if regra else "",
            regra.legislacao_ref or "" if regra else "",
            regra.observacoes or "" if regra else "",
        ])

    output = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=CCD_TTD_Completo.csv"},
    )


@router.get("/exportar/excel")
async def exportar_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exportar CCD/TTD em Excel (HU-033)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    arvore = await _carregar_arvore_completa(db)
    flat = _flatten_tree(arvore)
    regras_map = await _carregar_regras_por_nivel(db)
    holds = await _holds_ativos_por_nivel(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "CCD e TTD"

    # Estilos
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Código", "Código CONARQ", "Tipo", "Título", "Descrição",
        "Sigilo", "Status", "Evento Início",
        "Fase Corrente (anos)", "Fase Intermediária (anos)",
        "Destinação Final", "Base Legal", "Legislação Ref.", "Observações", "Hold Ativo",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, (nivel, _depth) in enumerate(flat, 2):
        regra = regras_map.get(nivel.id)
        tem_hold = nivel.id in holds
        vals = [
            nivel.codigo,
            nivel.codigo_conarq or "",
            LABEL_TIPO.get(nivel.tipo, nivel.tipo),
            nivel.titulo,
            nivel.descricao or "",
            nivel.nivel_sigilo,
            nivel.status,
            regra.evento_inicio if regra else "",
            regra.fase_corrente if regra else "",
            regra.fase_intermediaria if regra else "",
            LABEL_DESTINACAO.get(regra.destinacao_final, regra.destinacao_final) if regra else "",
            regra.base_legal or "" if regra else "",
            regra.legislacao_ref or "" if regra else "",
            regra.observacoes or "" if regra else "",
            "Sim" if tem_hold else "",
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
            if tem_hold:
                cell.fill = hold_fill

    # Ajustar largura
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 16
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=CCD_TTD_Completo.xlsx"},
    )


# ========== HU-034: Importação em Lote ==========

@router.post("/importar/excel", response_model=ImportResultResponse)
async def importar_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Importar níveis PCD + regras TTD a partir de planilha Excel (HU-034).

    Colunas esperadas (na ordem):
    Código | Código CONARQ | Tipo | Título | Descrição | Sigilo |
    Código Pai | Evento Início | Fase Corrente | Fase Intermediária |
    Destinação Final | Base Legal | Legislação Ref | Observações
    """
    from openpyxl import load_workbook

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo excede 10 MB")

    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    criados = 0
    erros: list[str] = []
    codigos_existentes: dict[str, uuid.UUID] = {}

    # Cache de códigos existentes
    existing = await db.execute(select(PcdNivel.codigo, PcdNivel.id))
    for row in existing.all():
        codigos_existentes[row[0]] = row[1]

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row_num, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue

        try:
            codigo = str(row[0]).strip()
            codigo_conarq = str(row[1]).strip() if row[1] else None
            tipo = str(row[2]).strip().lower() if row[2] else "classe"
            titulo = str(row[3]).strip() if row[3] else codigo
            descricao = str(row[4]).strip() if len(row) > 4 and row[4] else None
            sigilo = str(row[5]).strip().lower() if len(row) > 5 and row[5] else "publico"
            codigo_pai = str(row[6]).strip() if len(row) > 6 and row[6] else None
            evento_inicio = str(row[7]).strip() if len(row) > 7 and row[7] else None
            fase_corrente = int(row[8]) if len(row) > 8 and row[8] else 0
            fase_intermediaria = int(row[9]) if len(row) > 9 and row[9] else 0
            destinacao_final = str(row[10]).strip().lower() if len(row) > 10 and row[10] else None
            base_legal = str(row[11]).strip() if len(row) > 11 and row[11] else None
            legislacao_ref = str(row[12]).strip() if len(row) > 12 and row[12] else None
            observacoes = str(row[13]).strip() if len(row) > 13 and row[13] else None

            if tipo not in TIPOS_NIVEL_VALIDOS:
                erros.append(f"Linha {row_num}: tipo '{tipo}' inválido")
                continue

            if codigo in codigos_existentes:
                erros.append(f"Linha {row_num}: código '{codigo}' já existe")
                continue

            pai_id = None
            if codigo_pai:
                pai_id = codigos_existentes.get(codigo_pai)
                if not pai_id:
                    erros.append(f"Linha {row_num}: código pai '{codigo_pai}' não encontrado")
                    continue

            nivel = PcdNivel(
                pai_id=pai_id,
                tipo=tipo,
                codigo=codigo,
                titulo=titulo,
                descricao=descricao,
                codigo_conarq=codigo_conarq,
                nivel_sigilo=sigilo,
                metadados={},
                responsavel_id=current_user.id,
            )
            db.add(nivel)
            await db.flush()
            codigos_existentes[codigo] = nivel.id

            # Criar regra TTD se destinação informada
            if destinacao_final and destinacao_final in DESTINACOES_VALIDAS:
                regra = RegraRetencao(
                    pcd_nivel_id=nivel.id,
                    evento_inicio=evento_inicio or "criacao",
                    prazo_dias=(fase_corrente + fase_intermediaria) * 365,
                    fase_corrente=fase_corrente,
                    fase_intermediaria=fase_intermediaria,
                    destinacao_final=destinacao_final,
                    base_legal=base_legal,
                    legislacao_ref=legislacao_ref,
                    observacoes=observacoes,
                )
                db.add(regra)
                await db.flush()

            criados += 1
        except Exception as exc:
            erros.append(f"Linha {row_num}: {exc}")

    return ImportResultResponse(criados=criados, erros=erros)


# ========== HU-037: Listagem / Termo de Eliminação ==========

@router.get("/termo-eliminacao", response_model=TermoEliminacaoResponse)
async def gerar_termo_eliminacao(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Gerar Listagem de Eliminação de documentos com prazo cumprido (HU-037)."""
    regras = await db.execute(
        select(RegraRetencao).where(RegraRetencao.destinacao_final == "eliminacao")
    )
    regras_list = regras.scalars().all()
    holds = await _holds_ativos_por_nivel(db)

    nivel_ids = {r.pcd_nivel_id for r in regras_list if r.pcd_nivel_id not in holds}
    niveis_result = await db.execute(select(PcdNivel).where(PcdNivel.id.in_(nivel_ids)))
    niveis_map = {n.id: n for n in niveis_result.scalars().all()}

    itens: list[TermoEliminacaoItem] = []
    for regra in regras_list:
        if regra.pcd_nivel_id in holds:
            continue
        nivel = niveis_map.get(regra.pcd_nivel_id)
        if not nivel:
            continue
        itens.append(TermoEliminacaoItem(
            codigo=nivel.codigo,
            titulo=nivel.titulo,
            tipo=nivel.tipo,
            destinacao_final=regra.destinacao_final,
            base_legal=regra.base_legal,
            fase_corrente=regra.fase_corrente,
            fase_intermediaria=regra.fase_intermediaria,
            observacoes=regra.observacoes,
        ))

    return TermoEliminacaoResponse(
        titulo="Listagem de Eliminação de Documentos",
        data_geracao=datetime.now(UTC),
        gerado_por=current_user.email,
        itens=sorted(itens, key=lambda x: x.codigo),
        total_itens=len(itens),
    )


@router.get("/termo-eliminacao/pdf")
async def termo_eliminacao_pdf(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Gerar Termo de Eliminação em PDF (HU-037)."""
    termo = await gerar_termo_eliminacao(db=db, current_user=current_user)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "TERMO DE ELIMINACAO DE DOCUMENTOS", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Data: {termo.data_geracao.strftime('%d/%m/%Y %H:%M UTC')}  |  Responsavel: {termo.gerado_por}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(4)

    col_w = [25, 60, 25, 30, 20, 20, 30, 50]
    headers = ["Codigo", "Titulo", "Tipo", "Destinacao", "Corrente", "Intermed.", "Base Legal", "Obs."]

    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(200, 200, 200)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 6)
    for item in termo.itens:
        vals = [
            item.codigo,
            item.titulo[:50],
            LABEL_TIPO.get(item.tipo, item.tipo),
            LABEL_DESTINACAO.get(item.destinacao_final, item.destinacao_final),
            f"{item.fase_corrente}a",
            f"{item.fase_intermediaria}a",
            (item.base_legal or "")[:25],
            (item.observacoes or "")[:40],
        ]
        for i, v in enumerate(vals):
            pdf.cell(col_w[i], 5, v, border=1)
        pdf.ln()

    pdf.ln(8)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Total de itens para eliminacao: {termo.total_itens}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(15)
    pdf.cell(90, 6, "____________________________________", align="C")
    pdf.cell(90, 6, "____________________________________", align="C")
    pdf.ln()
    pdf.cell(90, 6, "Responsavel pela Eliminacao", align="C")
    pdf.cell(90, 6, "Autoridade Arquivistica", align="C")

    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Termo_Eliminacao.pdf"},
    )


# ========== HU-038: Relatório de Transferência / Recolhimento ==========

@router.get("/relatorio-transferencia", response_model=RelatorioTransferenciaResponse)
async def gerar_relatorio_transferencia(
    tipo: str = Query("transferencia", description="transferencia | recolhimento"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Gerar relatório de transferência/recolhimento (HU-038)."""
    dest_map = {
        "transferencia": "guarda_permanente",
        "recolhimento": "guarda_permanente",
    }
    if tipo not in dest_map:
        raise HTTPException(status_code=400, detail="Tipo deve ser 'transferencia' ou 'recolhimento'")

    regras = await db.execute(
        select(RegraRetencao).where(RegraRetencao.destinacao_final == dest_map[tipo])
    )
    regras_list = regras.scalars().all()
    holds = await _holds_ativos_por_nivel(db)

    nivel_ids = {r.pcd_nivel_id for r in regras_list if r.pcd_nivel_id not in holds}
    niveis_result = await db.execute(select(PcdNivel).where(PcdNivel.id.in_(nivel_ids)))
    niveis_map = {n.id: n for n in niveis_result.scalars().all()}

    itens: list[RelatorioTransferenciaItem] = []
    for regra in regras_list:
        if regra.pcd_nivel_id in holds:
            continue
        nivel = niveis_map.get(regra.pcd_nivel_id)
        if not nivel:
            continue
        itens.append(RelatorioTransferenciaItem(
            codigo=nivel.codigo,
            titulo=nivel.titulo,
            tipo=nivel.tipo,
            destinacao_final=regra.destinacao_final,
            fase_corrente=regra.fase_corrente,
            fase_intermediaria=regra.fase_intermediaria,
            base_legal=regra.base_legal,
        ))

    label = "Transferência" if tipo == "transferencia" else "Recolhimento"
    return RelatorioTransferenciaResponse(
        titulo=f"Relatório de {label} de Documentos",
        tipo_relatorio=tipo,
        data_geracao=datetime.now(UTC),
        gerado_por=current_user.email,
        itens=sorted(itens, key=lambda x: x.codigo),
        total_itens=len(itens),
    )


@router.get("/relatorio-transferencia/pdf")
async def relatorio_transferencia_pdf(
    tipo: str = Query("transferencia"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Gerar relatório de transferência/recolhimento em PDF (HU-038)."""
    rel = await gerar_relatorio_transferencia(tipo=tipo, db=db, current_user=current_user)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, rel.titulo.upper(), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Data: {rel.data_geracao.strftime('%d/%m/%Y %H:%M UTC')}  |  Responsavel: {rel.gerado_por}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(4)

    col_w = [30, 80, 30, 30, 25, 25, 50]
    headers = ["Codigo", "Titulo", "Tipo", "Destinacao", "Corrente", "Intermed.", "Base Legal"]

    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(200, 200, 200)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for item in rel.itens:
        vals = [
            item.codigo,
            item.titulo[:60],
            LABEL_TIPO.get(item.tipo, item.tipo),
            LABEL_DESTINACAO.get(item.destinacao_final, item.destinacao_final),
            f"{item.fase_corrente}a",
            f"{item.fase_intermediaria}a",
            (item.base_legal or "")[:40],
        ]
        for i, v in enumerate(vals):
            pdf.cell(col_w[i], 5, v, border=1)
        pdf.ln()

    pdf.ln(8)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Total de itens: {rel.total_itens}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(15)
    pdf.cell(90, 6, "____________________________________", align="C")
    pdf.cell(90, 6, "____________________________________", align="C")
    pdf.ln()
    label = "Transferencia" if tipo == "transferencia" else "Recolhimento"
    pdf.cell(90, 6, f"Responsavel pela {label}", align="C")
    pdf.cell(90, 6, "Arquivo Destinatario", align="C")

    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Relatorio_{label}.pdf"},
    )
